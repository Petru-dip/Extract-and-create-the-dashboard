
# import os
# import glob
# from datetime import datetime
# import pandas as pd
# from openpyxl import Workbook
# from openpyxl.styles import Font
# from openpyxl.utils import get_column_letter


# FOLDER_INPUT = r"Rezultate Teste"
# OUTPUT_EXCEL = "dashboard_final.xlsx"


# def parse_quiz_file(file_path: str):
#     with open(file_path, "r", encoding="utf-8-sig") as f:
#         lines = [line.strip() for line in f.readlines()]

#     # === METADATA ===
#     meta_line = lines[0]
#     parts = meta_line.split(",")

#     name = email = "UNKNOWN"
#     timestamp = datetime(2100, 1, 1)

#     for p in parts:
#         p = p.strip()
#         if p.lower().startswith("name:"):
#             name = p.split(":", 1)[1].strip()
#         elif p.lower().startswith("email:"):
#             email = p.split(":", 1)[1].strip()
#         elif p.lower().startswith("startdate:"):
#             dt = p.split(":", 1)[1].strip()
#             timestamp = datetime.strptime(dt, "%Y-%m-%d %H:%M:%S")

#     # === Citire tabel CSV și total ===
#     df = pd.read_csv(file_path, skiprows=1, encoding="utf-8-sig")

#     last = df.iloc[-1]
#     if str(last["NR_INTREBARE"]).upper().startswith("TOTAL"):
#         total = int(last["PUNCTAJ"])
#         df_questions = df.iloc[:-1].copy()
#     else:
#         total = int(df["PUNCTAJ"].sum())
#         df_questions = df.copy()

#     return {
#         "name": name,
#         "email": email,
#         "timestamp": timestamp,
#         "total": total,
#         "questions_df": df_questions,
#     }


# def make_safe_sheet_name(name: str, existing: set) -> str:
#     safe = "".join(c for c in name if c.isalnum() or c in (" ", "_")).strip()
#     if not safe:
#         safe = "Sheet"
#     base = safe[:31]
#     out = base
#     i = 2
#     while out in existing:
#         suffix = f"_{i}"
#         out = base[: 31 - len(suffix)] + suffix
#         i += 1
#     existing.add(out)
#     return out


# def main():
#     csv_files = glob.glob(os.path.join(FOLDER_INPUT, "*.csv"))
#     if not csv_files:
#         print("Nu există fișiere în folder!")
#         return

#     records = []
#     persons = []
#     sheet_names = set()

#     # Procesăm fișierele
#     for file in csv_files:
#         data = parse_quiz_file(file)

#         sheet_name = make_safe_sheet_name(data["name"], sheet_names)

#         records.append({
#             "Name": data["name"],
#             "Email": data["email"],
#             "Total": data["total"],
#             "Timestamp": data["timestamp"],
#             "SheetName": sheet_name
#         })

#         persons.append({
#             "sheet_name": sheet_name,
#             "name": data["name"],
#             "email": data["email"],
#             "timestamp": data["timestamp"],
#             "total": data["total"],
#             "df": data["questions_df"]
#         })

#     # Dashboard
#     df_dash = pd.DataFrame(records).sort_values("Total", ascending=False)

#     wb = Workbook()
#     ws_dash = wb.active
#     ws_dash.title = "Dashboard"

#     # Header Dashboard
#     headers = ["Name", "Email", "Total", "Timestamp"]
#     ws_dash.append(headers)
#     for c in ws_dash[1]:
#         c.font = Font(bold=True)

#     # Rows Dashboard
#     for _, r in df_dash.iterrows():
#         ws_dash.append([r["Name"], r["Email"], r["Total"], r["Timestamp"]])

#     ws_dash.freeze_panes = "A2"

#     name_to_sheet = {r["Name"]: r["SheetName"] for _, r in df_dash.iterrows()}

#     # Creare sheet-uri individuale
#     for p in persons:
#         ws = wb.create_sheet(title=p["sheet_name"])

#         # === META în două coloane ===
#         ws.append(["Name", p["name"]])
#         ws.append(["Email", p["email"]])
#         ws.append(["StartDate", p["timestamp"].strftime("%Y-%m-%d %H:%M:%S")])
#         ws.append(["TOTAL", p["total"]])
#         ws.append([""])  # linie goală

#         # Bold pentru titluri
#         ws["A1"].font = Font(bold=True)
#         ws["A2"].font = Font(bold=True)
#         ws["A3"].font = Font(bold=True)
#         ws["A4"].font = Font(bold=True)

#         # === BUTON "Back to Dashboard" ===
#         ws["A6"] = "⬅ Back to Dashboard"
#         ws["A6"].hyperlink = "#'Dashboard'!A1"
#         ws["A6"].style = "Hyperlink"
#         ws["A6"].font = Font(color="0000FF", bold=True)

#         ws.append([""])  # spațiu

#         # === Header tabel ===
#         start_row = ws.max_row + 1
#         ws.append(list(p["df"].columns))
#         for c in ws[start_row]:
#             c.font = Font(bold=True)

#         # === Populare tabel ===
#         for _, row in p["df"].iterrows():
#             ws.append(row.tolist())

#         ws.freeze_panes = f"A{start_row+1}"

#     # Hyperlinkuri în dashboard
#     for row in range(2, ws_dash.max_row + 1):
#         name = ws_dash.cell(row=row, column=1).value
#         sheet_name = name_to_sheet.get(name)
#         if sheet_name:
#             cell = ws_dash.cell(row=row, column=1)
#             cell.hyperlink = f"#'{sheet_name}'!A1"
#             cell.style = "Hyperlink"

#     wb.save(OUTPUT_EXCEL)
#     print("✔️ Dashboard generat complet.")


# if __name__ == "__main__":
#     main()
import os
import glob
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference

FOLDER_INPUT = r"Rezultate Teste"
OUTPUT_EXCEL = "dashboard_final.xlsx"



def parse_quiz_file(file_path: str):
    """Citește fișierul CSV și extrage meta + întrebări + total."""
    with open(file_path, "r", encoding="utf-8-sig") as f:
        lines = [line.strip() for line in f.readlines()]

    # META
    meta_line = lines[0]
    parts = meta_line.split(",")

    name = email = "UNKNOWN"
    timestamp = datetime(2100, 1, 1)

    for p in parts:
        p = p.strip()
        if p.lower().startswith("name:"):
            name = p.split(":", 1)[1].strip()
        elif p.lower().startswith("email:"):
            email = p.split(":", 1)[1].strip()
        elif p.lower().startswith("startdate:"):
            dt = p.split(":", 1)[1].strip()
            timestamp = datetime.strptime(dt, "%Y-%m-%d %H:%M:%S")

    # TABEL + TOTAL
    df = pd.read_csv(file_path, skiprows=1, encoding="utf-8-sig")

    last = df.iloc[-1]
    if str(last["NR_INTREBARE"]).upper().startswith("TOTAL"):
        total = int(last["PUNCTAJ"])
        df_questions = df.iloc[:-1].copy()
    else:
        total = int(df["PUNCTAJ"].sum())
        df_questions = df.copy()

    return {
        "name": name,
        "email": email,
        "timestamp": timestamp,
        "total": total,
        "questions_df": df_questions,
    }



def make_safe_sheet_name(name: str, existing: set) -> str:
    safe = "".join(c for c in name if c.isalnum() or c in (" ", "_")).strip()
    if not safe:
        safe = "Sheet"

    base = safe[:31]
    out = base
    i = 2
    while out in existing:
        suffix = f"_{i}"
        out = base[: 31 - len(suffix)] + suffix
        i += 1

    existing.add(out)
    return out



def main():
    csv_files = glob.glob(os.path.join(FOLDER_INPUT, "*.csv"))
    if not csv_files:
        print("❌ Nu există fișiere!")
        return

    records = []
    persons = []
    existing_names = set()

    for file in csv_files:
        data = parse_quiz_file(file)

        sheet_name = make_safe_sheet_name(data["name"], existing_names)

        records.append({
            "Name": data["name"],
            "Email": data["email"],
            "Total": data["total"],
            "Timestamp": data["timestamp"],
            "SheetName": sheet_name
        })

        persons.append({
            "sheet_name": sheet_name,
            "name": data["name"],
            "email": data["email"],
            "timestamp": data["timestamp"],
            "total": data["total"],
            "df": data["questions_df"]
        })

    # ===== CREARE DASHBOARD =====
    df_dash = pd.DataFrame(records).sort_values("Total", ascending=False)

    wb = Workbook()
    ws_dash = wb.active
    ws_dash.title = "Dashboard"

    # Header
    headers = ["Name", "Email", "Total", "Timestamp"]
    ws_dash.append(headers)

    for cell in ws_dash[1]:
        cell.font = Font(bold=True)

    # Rows
    for _, r in df_dash.iterrows():
        ws_dash.append([r["Name"], r["Email"], r["Total"], r["Timestamp"]])

    ws_dash.freeze_panes = "A2"

    # ===== Total participanți =====
    total_part = len(df_dash)
    ws_dash["F1"] = f"Total participanți: {total_part}"
    ws_dash["F1"].font = Font(bold=True)

    # ===== HISTOGRAMĂ (grafic distribuție scoruri) =====
    # Construim tabel cu intervale
    intervals = ["0–5", "6–10", "11–15", "16–20", "21–25", "26–30", "31–35", "36–40"]
    bins = [(0,5),(6,10),(11,15),(16,20),(21,25),(26,30),(31,35),(36,40)]

    ws_dash.append([])  # blank
    start_row = ws_dash.max_row + 1

    ws_dash[f"G{start_row}"] = "Interval"
    ws_dash[f"H{start_row}"] = "Număr"
    ws_dash[f"G{start_row}"].font = Font(bold=True)
    ws_dash[f"H{start_row}"].font = Font(bold=True)

    # Umple tabelul
    counts = []
    for label, (lo, hi) in zip(intervals, bins):
        count = ((df_dash["Total"] >= lo) & (df_dash["Total"] <= hi)).sum()
        counts.append((label, count))

    for i, (label, val) in enumerate(counts, start=start_row+1):
        ws_dash[f"G{i}"] = label
        ws_dash[f"H{i}"] = val

    # Creăm graficul
    chart = BarChart()
    chart.title = "Distribuția scorurilor"
    chart.x_axis.title = "Interval scor"
    chart.y_axis.title = "Număr participanți"

    data_ref = Reference(ws_dash, min_col=8, min_row=start_row, max_row=start_row+len(bins))
    cats_ref = Reference(ws_dash, min_col=7, min_row=start_row+1, max_row=start_row+len(bins))

    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    ws_dash.add_chart(chart, f"J{start_row}")

    # ===== SHEET-URI INDIVIDUALE =====
    name_to_sheet = {r["Name"]: r["SheetName"] for _, r in df_dash.iterrows()}

    for p in persons:
        ws = wb.create_sheet(title=p["sheet_name"])

        # META în două coloane
        ws.append(["Name", p["name"]])
        ws.append(["Email", p["email"]])
        ws.append(["StartDate", p["timestamp"].strftime("%Y-%m-%d %H:%M:%S")])
        ws.append(["TOTAL", p["total"]])
        ws.append([""])  # blank

        ws["A1"].font = ws["A2"].font = ws["A3"].font = ws["A4"].font = Font(bold=True)

        # Buton înapoi
        ws["A6"] = "⬅ Back to Dashboard"
        ws["A6"].hyperlink = "#'Dashboard'!A1"
        ws["A6"].style = "Hyperlink"
        ws["A6"].font = Font(color="0000FF", bold=True)
        ws.append([""])

        # Header tabel întrebări
        start = ws.max_row + 1
        ws.append(list(p["df"].columns))
        for c in ws[start]:
            c.font = Font(bold=True)

        for _, row in p["df"].iterrows():
            ws.append(row.tolist())

        ws.freeze_panes = f"A{start+1}"

    # Hyperlink în dashboard
    for row in range(2, ws_dash.max_row + 1):
        name = ws_dash.cell(row=row, column=1).value
        target = name_to_sheet.get(name)
        if target:
            cell = ws_dash.cell(row=row, column=1)
            cell.hyperlink = f"#'{target}'!A1"
            cell.style = "Hyperlink"

    wb.save(OUTPUT_EXCEL)
    print("✔️ Dashboard complet generat cu succes!")


if __name__ == "__main__":
    main()
